"""Document ingestion module for parsing various file formats."""

from pathlib import Path
from typing import List, Dict, Optional, Union
import hashlib
import mimetypes
import re

import PyPDF2
from docx import Document
from openpyxl import load_workbook

from .config import OUTPUT_DIR


MAX_NATIVE_PDF_BYTES = 32 * 1024 * 1024  # 32 MB
MAX_NATIVE_PDF_PAGES = 100
PDF_CHUNK_PAGE_LIMIT = 20
PDF_CHUNK_OVERLAP = 2
PDF_CHUNK_DIR = OUTPUT_DIR / "artifacts" / "pdf_chunks"
PDF_CHUNK_DIR.mkdir(parents=True, exist_ok=True)
SUPPORTED_IMAGE_FORMATS = {
    '.png', '.jpg', '.jpeg', '.gif', '.webp', '.tif', '.tiff'
}


class DocumentIngester:
    """Handles reading and parsing of various document formats."""

    def __init__(self):
        self.supported_formats = {'.pdf', '.txt', '.md', '.vtt', '.docx', '.xlsx'}
        # Filenames to ignore in input directories (case-insensitive)
        self.ignored_filenames = {'readme.txt'}
    
    def ingest_directory(self, directory: Path) -> List[Dict[str, str]]:
        """
        Ingest all supported documents from a directory.
        
        Args:
            directory: Path to directory containing documents
            
        Returns:
            List of dictionaries with 'filename' and 'content' keys
        """
        documents = []
        
        if not directory.exists():
            print(f"[WARN] Directory {directory} does not exist")
            return documents
        
        for file_path in directory.iterdir():
            if not file_path.is_file():
                continue

            suffix = file_path.suffix.lower()

            if file_path.name.lower() in self.ignored_filenames:
                print(f"[INFO] Skipping {file_path.name} (ignored)")
                continue

            if suffix in self.supported_formats:
                document = self.ingest_file(file_path)
            elif suffix in SUPPORTED_IMAGE_FORMATS:
                document = self.ingest_image(file_path)
            else:
                print(f"[WARN] Unsupported file format {suffix} for {file_path.name}")
                continue

            if not document:
                continue

            if isinstance(document, list):
                documents.extend(document)
            else:
                documents.append(document)
            print(f"[OK] Ingested: {file_path.name}")
        
        return documents
    
    def ingest_file(self, file_path: Path) -> Optional[Union[Dict[str, str], List[Dict[str, str]]]]:
        """
        Ingest a single file based on its format.
        
        Args:
            file_path: Path to the file
            
        Returns:
            Document dictionary with content and metadata
        """
        suffix = file_path.suffix.lower()
        
        try:
            if suffix == '.pdf':
                return self._process_pdf(file_path)
            elif suffix in ['.txt', '.md']:
                content = self._read_text(file_path)
                is_summary = file_path.name.endswith('.summary.txt')
                source_type = 'summary' if is_summary else 'text'
                upload_via = 'summary' if is_summary else 'text'
                original_filename = file_path.name[:-len('.summary.txt')] if is_summary else file_path.name
                return {
                    'filename': file_path.name,
                    'content': content,
                    'path': str(file_path),
                    'media_type': mimetypes.guess_type(file_path.name)[0] or 'text/plain',
                    'source_type': source_type,
                    'size_bytes': file_path.stat().st_size,
                    'upload_via': upload_via,
                    'can_upload': False,
                    'content_hash': self._hash_text(content),
                    'metadata': {
                        'original_filename': original_filename,
                        'summary_mode': is_summary,
                    }
                }
            elif suffix == '.vtt':
                content = self._read_vtt(file_path)
                return {
                    'filename': file_path.name,
                    'content': content,
                    'path': str(file_path),
                    'media_type': 'text/vtt',
                    'source_type': 'transcript',
                    'size_bytes': file_path.stat().st_size,
                    'upload_via': 'text',
                    'can_upload': False,
                    'content_hash': self._hash_text(content),
                }
            elif suffix == '.docx':
                return self._process_docx(file_path)
            elif suffix == '.xlsx':
                return self._process_xlsx(file_path)
            else:
                print(f"[WARN] Unsupported file format {suffix}")
                return None
        except Exception as e:
            print(f"[ERROR] Error reading {file_path.name}: {str(e)}")
            return None

    def ingest_image(self, file_path: Path) -> Optional[Dict[str, str]]:
        """Prepare an image file for downstream processing."""

        size_bytes = file_path.stat().st_size
        media_type = mimetypes.guess_type(file_path.name)[0] or 'image/png'
        can_upload = size_bytes <= MAX_NATIVE_PDF_BYTES  # reuse 32 MB limit

        if not can_upload:
            print(f"[WARN] Image {file_path.name} exceeds upload limits ({size_bytes} bytes)")

        placeholder = f"[IMAGE] {file_path.name}"

        return {
            'filename': file_path.name,
            'content': placeholder,
            'path': str(file_path),
            'media_type': media_type,
            'source_type': 'image',
            'size_bytes': size_bytes,
            'upload_via': 'attachment' if can_upload else 'skipped',
            'can_upload': can_upload,
            'content_hash': self._hash_file(file_path),
        }

    def _process_pdf(self, file_path: Path) -> Union[Dict[str, str], List[Dict[str, str]]]:
        """Process a PDF, choosing between native upload, text extraction, and page chunking."""

        size_bytes = file_path.stat().st_size
        can_upload = size_bytes <= MAX_NATIVE_PDF_BYTES

        with open(file_path, 'rb') as file:
            pdf_reader = PyPDF2.PdfReader(file)
            page_count = len(pdf_reader.pages)
            page_texts: List[str] = []
            for page in pdf_reader.pages:
                page_text = page.extract_text() or ""
                page_texts.append(page_text.strip())

        if page_count > MAX_NATIVE_PDF_PAGES:
            can_upload = False

        combined_text = []
        for idx, text in enumerate(page_texts, start=1):
            if text:
                combined_text.append(f"--- Page {idx} ---\n{text}")

        text_content = "\n\n".join(combined_text).strip()

        if can_upload:
            if not text_content:
                text_content = (
                    f"[PDF] {file_path.name} (page count: {page_count or 'unknown'}) provided via native upload."
                )

            return {
                'filename': file_path.name,
                'content': text_content or f"[PDF] {file_path.name}",
                'path': str(file_path),
                'media_type': 'application/pdf',
                'source_type': 'pdf',
                'size_bytes': size_bytes,
                'page_count': page_count or 0,
                'upload_via': 'attachment',
                'can_upload': True,
                'content_hash': self._hash_file(file_path),
            }

        # Large PDF: rely on extracted text if available, otherwise split into chunks
        if text_content:
            return {
                'filename': file_path.name,
                'content': text_content,
                'path': str(file_path),
                'media_type': 'application/pdf',
                'source_type': 'pdf',
                'size_bytes': size_bytes,
                'page_count': page_count or 0,
                'upload_via': 'text',
                'can_upload': False,
                'content_hash': self._hash_file(file_path),
            }

        return self._split_pdf_into_chunks(file_path, page_texts)

    def _split_pdf_into_chunks(self, file_path: Path, page_texts: List[str]) -> List[Dict[str, str]]:
        chunks: List[Dict[str, str]] = []
        chunk_index = 0
        total_pages = len(page_texts)

        while chunk_index < total_pages:
            start = chunk_index
            end = min(start + PDF_CHUNK_PAGE_LIMIT, total_pages)
            pages = list(range(start, end))

            # Include overlap from previous chunk
            if chunk_index > 0 and PDF_CHUNK_OVERLAP > 0:
                overlap_start = max(0, start - PDF_CHUNK_OVERLAP)
                previous_pages = list(range(overlap_start, start))
                pages = previous_pages + pages

            chunk_texts = []
            for page_idx in pages:
                page_number = page_idx + 1
                text = page_texts[page_idx]
                if text:
                    chunk_texts.append(f"--- Page {page_number} ---\n{text}")

            # Skip empty chunks (e.g., blank pages)
            if not chunk_texts:
                chunk_index += PDF_CHUNK_PAGE_LIMIT
                continue

            chunk_body = "\n\n".join(chunk_texts)
            chunk_filename = f"{file_path.stem}_chunk_{start+1:04d}_{min(end, total_pages):04d}.pdf.txt"
            chunk_path = PDF_CHUNK_DIR / chunk_filename
            with open(chunk_path, 'w', encoding='utf-8') as f:
                f.write(chunk_body)

            chunks.append({
                'filename': chunk_filename,
                'content': chunk_body,
                'path': str(chunk_path),
                'media_type': 'application/pdf',
                'source_type': 'pdf_chunk',
                'size_bytes': len(chunk_body.encode('utf-8', errors='ignore')),
                'page_count': len(pages),
                'upload_via': 'chunk',
                'can_upload': False,
                'content_hash': self._hash_text(chunk_body),
                'parent_file': file_path.name,
                'page_range': {'start': pages[0] + 1, 'end': pages[-1] + 1},
            })

            chunk_index += PDF_CHUNK_PAGE_LIMIT

        return chunks

    def _process_docx(self, file_path: Path) -> Dict[str, str]:
        doc = Document(file_path)
        paragraphs = [p.text.strip() for p in doc.paragraphs if p.text.strip()]
        content = "\n".join(paragraphs)
        media_type = 'application/vnd.openxmlformats-officedocument.wordprocessingml.document'
        size_bytes = file_path.stat().st_size
        can_upload = size_bytes <= MAX_NATIVE_PDF_BYTES

        if not content:
            content = f"[DOCX] {file_path.name}"

        return {
            'filename': file_path.name,
            'content': content,
            'path': str(file_path),
            'media_type': media_type,
            'source_type': 'docx',
            'size_bytes': size_bytes,
            'upload_via': 'attachment' if can_upload else 'text',
            'can_upload': can_upload,
            'content_hash': self._hash_file(file_path),
        }

    def _process_xlsx(self, file_path: Path) -> Dict[str, str]:
        """Extract a text summary from an Excel workbook (no native upload)."""
        try:
            wb = load_workbook(filename=str(file_path), data_only=True, read_only=True)
        except Exception as exc:
            print(f"[WARN] Could not open XLSX ({file_path.name}): {exc}")
            return {
                'filename': file_path.name,
                'content': f"[XLSX] {file_path.name} (unreadable)",
                'path': str(file_path),
                'media_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                'source_type': 'xlsx',
                'size_bytes': file_path.stat().st_size,
                'upload_via': 'text',
                'can_upload': False,
                'content_hash': self._hash_file(file_path),
            }

        lines: List[str] = []
        sheet_limit = 10
        row_limit = 200
        col_limit = 20
        try:
            for si, sheet in enumerate(wb.worksheets[:sheet_limit], start=1):
                lines.append(f"--- Sheet {si}: {sheet.title} ---")
                rows = 0
                for r in sheet.iter_rows(min_row=1, max_row=row_limit, max_col=col_limit, values_only=True):
                    rows += 1
                    vals = ["" if v is None else str(v) for v in r]
                    lines.append("\t".join(vals))
                if sheet.max_row > row_limit or sheet.max_column > col_limit:
                    lines.append("[... truncated ...]")
        except Exception as exc:
            lines.append(f"[WARN] Error reading sheet data: {exc}")
        finally:
            try:
                wb.close()
            except Exception:
                pass

        content = "\n".join(lines) if lines else f"[XLSX] {file_path.name} (empty)"
        return {
            'filename': file_path.name,
            'content': content,
            'path': str(file_path),
            'media_type': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            'source_type': 'xlsx',
            'size_bytes': file_path.stat().st_size,
            'upload_via': 'text',  # API does not accept XLSX as native document
            'can_upload': False,
            'content_hash': self._hash_file(file_path),
        }
    
    def _read_text(self, file_path: Path) -> str:
        """Read plain text or markdown file."""
        with open(file_path, 'r', encoding='utf-8') as file:
            return file.read()

    def _hash_text(self, text: str) -> str:
        hasher = hashlib.sha256()
        hasher.update(text.encode('utf-8', errors='ignore'))
        return hasher.hexdigest()

    def _hash_file(self, file_path: Path) -> str:
        hasher = hashlib.sha256()
        with open(file_path, 'rb') as f:
            for chunk in iter(lambda: f.read(8192), b""):
                hasher.update(chunk)
        return hasher.hexdigest()

    def _read_vtt(self, file_path: Path) -> str:
        """Parse a WebVTT transcript, removing timestamps and cue indices.

        Preserves speaker labels if present (e.g., "SPEAKER: text").
        """
        with open(file_path, 'r', encoding='utf-8') as f:
            lines = f.read().splitlines()

        cleaned_lines: List[str] = []

        timestamp_pattern = re.compile(r"\d{2}:\d{2}:\d{2}\.\d{3}\s+-->\s+\d{2}:\d{2}:\d{2}\.\d{3}")

        for line in lines:
            if not line:
                continue
            if line.strip().upper().startswith('WEBVTT'):
                continue
            # Skip cue numbers (lines that are only digits)
            if line.strip().isdigit():
                continue
            # Skip timestamp lines
            if timestamp_pattern.search(line):
                continue
            # Remove common vtt cue settings like "align:start position:0%"
            if 'align:' in line or 'position:' in line or 'line:' in line or 'size:' in line:
                # Often appear on the same line as text; keep text after a space if any
                parts = re.split(r"\s+(align:|position:|line:|size:)", line, maxsplit=1)
                line = parts[0].strip()
                if not line:
                    continue
            cleaned_lines.append(line.strip())

        # Merge short caption fragments by joining lines, but keep paragraph breaks
        text = " ".join(cleaned_lines)
        # Collapse multiple spaces
        text = re.sub(r"\s+", " ", text).strip()
        return text
    
    def combine_documents(self, documents: List[Dict[str, str]]) -> str:
        """
        Combine multiple documents into a single text block.
        
        Args:
            documents: List of document dictionaries
            
        Returns:
            Combined text with document separators
        """
        combined = []
        
        for doc in documents:
            separator = f"\n\n{'='*80}\n"
            separator += f"DOCUMENT: {doc['filename']}\n"
            separator += f"{'='*80}\n\n"
            meta_lines = [
                f"Source type: {doc.get('source_type', 'unknown')}",
                f"Ingest method: {doc.get('upload_via', 'text')}"
            ]
            if doc.get('page_count'):
                meta_lines.append(f"Page count: {doc['page_count']}")
            if doc.get('size_bytes'):
                meta_lines.append(f"Size: {doc['size_bytes']} bytes")
            meta_block = "\n".join(meta_lines)
            combined.append(separator + meta_block + "\n\n" + doc['content'])
        
        return "\n\n".join(combined)

